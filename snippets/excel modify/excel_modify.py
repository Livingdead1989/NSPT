import openpyxl

workbook = openpyxl.load_workbook('transactions.xlsx')
sheet = workbook['Sheet1']

#cell = sheet['a1']
#cell = sheet.cell(1, 1)
#print(cell.value)
#print(sheet.max_row)


for row in range(2, sheet.max_row + 1): # avoids headers & +1 for all content
    cell = sheet.cell(row, 3)
    corrected_value = cell.value * 0.9

    corrected_value_cell = sheet.cell(row, 4)
    corrected_value_cell.value = corrected_value

workbook.save('transactions_corrected.xlsx')